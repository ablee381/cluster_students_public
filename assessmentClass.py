import numpy as np
import openpyxl
import pandas as pd
from xls2xlsx import XLS2XLSX


class Results:
    def __init__(self, resultSheet, qNum=0, guessRate=0.25):
        # check if the spreadsheet is .xls
        fileParts = resultSheet.split('.')
        fileName = fileParts[0]
        xl_format = fileParts[1]
        if xl_format == 'xls':
            x = XLS2XLSX(resultSheet)
            wb = x.to_xlsx()
            resultSheet = fileName + '.xlsx'
            wb.save(resultSheet)
        self.resultSheet = resultSheet
        self.qNum = qNum
        self.guessRate = guessRate
        self.cutoff = qNum * guessRate + np.sqrt(qNum * guessRate * (1 - guessRate))
        wb_obj = openpyxl.load_workbook(self.resultSheet)
        sheet_obj = wb_obj.active
        cell_obj = sheet_obj.cell(1, 1)
        if not cell_obj.value:
            self.type = 'Potts'
        elif cell_obj.value == 'Timestamp':
            self.type = 'Google Forms'
        elif cell_obj.value == 'Local Student Id':
            self.type = 'Illuminate'
        else:
            self.type = 'Progress Learning'
        self.data = None

    def updateCutoff(self):
        self.cutoff = self.qNum * self.guessRate + np.sqrt(self.qNum * self.guessRate * (1 - self.guessRate))

    def read_norm_progress(self):
        self.data = pd.read_excel(self.resultSheet, header=2)
        self.data.dropna(axis=0, how='any', inplace=True)
        self.updateCutoff()
        header = list(self.data.columns)
        self.data = self.data.drop(self.data[self.data[header[1]] < self.cutoff].index)
        # get just the standards
        standards = []
        for i in range(len(header)):
            if 'MGSE' in header[i]:
                standards.append(header[i])
        names = self.data['Student']
        self.data = self.data.iloc[1:, :][standards]
        # Normalize the data to be 0 to 1
        self.data = (self.data - self.data.min()) / (self.data.max() - self.data.min())
        # Reinsert the names
        self.data.insert(0, 'name', names)

    def read_potts(self):
        """Reads and grades potts result spreadsheet
        Updates the result object's data field"""
        d = {}
        wb = openpyxl.load_workbook(self.resultSheet)
        sheet_obj = wb.active
        m_row = sheet_obj.max_row
        m_col = sheet_obj.max_column
        # get the student names
        namesL = []
        for i in range(8, m_row - 1):
            name = sheet_obj.cell(row=i, column=2).value
            namesL.append(name)
        d['name'] = namesL
        for j in range(8, m_col + 1):
            key = sheet_obj.cell(row=3, column=j).value
            if not key:
                continue
            resultL = []
            for i in range(8, m_row - 1):
                c = sheet_obj.cell(row=i, column=j)
                # sheet has student answer only if wrong, blank if correct
                if not c.value:
                    resultL.append(1)
                else:
                    resultL.append(0)
            d[key] = resultL
        self.data = pd.DataFrame(d)

    def read_illuminate(self):
        '''Reads and grades illuminate result spreadsheet
        Updates the result object's data field'''
        d = {}
        wb = openpyxl.load_workbook(self.resultSheet)
        sheet_obj = wb.active
        m_row = sheet_obj.max_row
        m_col = sheet_obj.max_column
        # get the student names
        namesL = []
        for i in range(2, m_row + 1):
            last = str(sheet_obj.cell(row=i, column=2).value)
            first = str(sheet_obj.cell(row=i, column=3).value)
            name = last + ', ' + first
            namesL.append(name)
        d['name'] = namesL
        for j in range(10, m_col + 1):
            key = sheet_obj.cell(row=1, column=j).value
            resultL = []
            for i in range(2, m_row + 1):
                c = sheet_obj.cell(row=i, column=j)
                fontColor = c.font.color.index
                if fontColor == 1 or fontColor == '00000000':
                    resultL.append(1)
                else:  # color.index='FFFF0000'
                    resultL.append(0)
            d[key] = resultL
        self.data = pd.DataFrame(d)

    def norm_illuminate(self):
        '''Removes any tests that do not make the cutoff; results are already scaled'''

        header = list(self.data.columns)
        totalNum = len(header) - 1

        self.qNum = totalNum
        self.updateCutoff()

        scores = self.data.sum(axis=1, numeric_only=True).array
        self.data = self.data.drop(self.data[scores < self.cutoff].index)

    def add_result(self, other):
        ''' adds the data from other to the calling result obj.
        throws out any Nan rows caused by kids not making up their tests '''
        self.data.join(other.data.set_index('Student'), on='Student')
        self.data.dropna(inplace=True)


# results_obj = Results('LSHS_Algebra_I_Unit_2_Spring_2023_Student_Responses-333509.xls')
